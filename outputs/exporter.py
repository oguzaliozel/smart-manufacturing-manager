import os
import sqlite3
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── TÜRKÇE FONT YÜKLEME (Windows Arial / Fallback Helvetica) ────────────────
FONT_NAME = "Helvetica"
FONT_BOLD_NAME = "Helvetica-Bold"

# Windows üzerindeki Arial yazı tipini Türkçe karakter desteği için kaydediyoruz
arial_path = "C:\\Windows\\Fonts\\arial.ttf"
arial_bold_path = "C:\\Windows\\Fonts\\arialbd.ttf"

if os.path.exists(arial_path) and os.path.exists(arial_bold_path):
    try:
        pdfmetrics.registerFont(TTFont("Arial", arial_path))
        pdfmetrics.registerFont(TTFont("Arial-Bold", arial_bold_path))
        FONT_NAME = "Arial"
        FONT_BOLD_NAME = "Arial-Bold"
    except Exception as e:
        print("Türkçe yazı tipi yükleme hatası:", e)


def export_to_excel(teklif, kalemler, filename):
    """
    Teklifi detaylı bir biçimde stilize edilmiş Excel dosyası olarak kaydeder.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teklif Detayı"
    ws.views.sheetView[0].showGridLines = True
    
    # Yazı Tipleri ve Renkler
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1E3A8A")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="6B7280")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True, color="111827")
    regular_font = Font(name=font_family, size=11, color="374151")
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    summary_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    total_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    scrap_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='111827'),
        bottom=Side(style='double', color='111827')
    )
    
    # Başlık Alanı
    ws.merge_cells("A1:G1")
    ws["A1"] = "ATÖLYE YÖNETİM SİSTEMİ"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("A2:G2")
    ws["A2"] = "Resmi Teklif Formu ve Maliyet Analizi Dökümü"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    
    # Metadata Blokları
    metadata = [
        ("Teklif No:", teklif["teklif_no"], "Oluşturma Tarihi:", teklif["olusturma_tarihi"]),
        ("Teklif Başlığı:", teklif["baslik"], "Teslimat Tarihi:", teklif["teslim_tarihi"] or "-"),
        ("Müşteri Firma:", teklif["firma_adi"] or "-", "Teklif Durumu:", teklif["durum"]),
        ("Müşteri Tel:", teklif["telefon"] or "-", "Müşteri E-Posta:", teklif["mail"] or "-")
    ]
    
    row_idx = 4
    for item in metadata:
        ws.cell(row=row_idx, column=1, value=item[0]).font = bold_font
        ws.cell(row=row_idx, column=2, value=item[1]).font = regular_font
        ws.cell(row=row_idx, column=4, value=item[2]).font = bold_font
        ws.cell(row=row_idx, column=5, value=item[3]).font = regular_font
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    row_idx += 1 # Boşluk
    
    # Tablo Başlıkları
    headers = ["Malzeme Açıklaması", "Miktar", "Birim", "Birim Fiyat", "Toplam Tutar", "Fire Miktarı", "Hurda Değeri"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    ws.row_dimensions[row_idx].height = 25
    row_idx += 1
    
    # Tablo Satırları
    for k in kalemler:
        ws.cell(row=row_idx, column=1, value=k["malzeme_adi"]).font = regular_font
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        
        c_qty = ws.cell(row=row_idx, column=2, value=k["miktar"])
        c_qty.font = regular_font
        c_qty.number_format = '#,##0.00'
        c_qty.alignment = Alignment(horizontal="right", vertical="center")
        
        ws.cell(row=row_idx, column=3, value=k["birim"]).font = regular_font
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
        
        c_price = ws.cell(row=row_idx, column=4, value=k["birim_fiyat"])
        c_price.font = regular_font
        c_price.number_format = '#,##0.00 "₺"'
        c_price.alignment = Alignment(horizontal="right", vertical="center")
        
        c_tot = ws.cell(row=row_idx, column=5, value=k["kalem_maliyeti"])
        c_tot.font = regular_font
        c_tot.number_format = '#,##0.00 "₺"'
        c_tot.alignment = Alignment(horizontal="right", vertical="center")
        
        c_fire = ws.cell(row=row_idx, column=6, value=k["fire_miktari"])
        c_fire.font = regular_font
        c_fire.number_format = '#,##0.00'
        c_fire.alignment = Alignment(horizontal="right", vertical="center")
        
        c_scrap = ws.cell(row=row_idx, column=7, value=k["tahmini_hurda_degeri"])
        c_scrap.font = regular_font
        c_scrap.number_format = '#,##0.00 "₺"'
        c_scrap.alignment = Alignment(horizontal="right", vertical="center")
        
        for c in range(1, 8):
            ws.cell(row=row_idx, column=c).border = thin_border
            
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        
    row_idx += 1 # Boşluk
    
    # Özet Satırları
    summary_data = [
        ("Malzeme Toplam Maliyeti:", teklif["malzeme_maliyeti"]),
        ("Makine Toplam Maliyeti:", teklif["makine_maliyeti"]),
        ("Ek Operasyonel Giderler:", teklif["ek_gider"]),
        ("Net Üretim Maliyeti:", teklif["net_maliyet"]),
        ("Planlanan Kar Oranı / Tutarı:", teklif["kar_tutari"]),
        ("Manuel Uygulanan İndirim:", -teklif["manuel_indirim"]),
        ("Son Teklif Tutarı (KDV Hariç):", teklif["son_tutar"]),
        ("Geri Kazanılabilir Hurda Değeri:", teklif["tahmini_hurda_degeri"])
    ]
    
    for label, val in summary_data:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        lbl_cell = ws.cell(row=row_idx, column=1, value=label)
        lbl_cell.font = bold_font if "Net" in label or "Son" in label or "Hurda" in label else regular_font
        lbl_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        val_cell = ws.cell(row=row_idx, column=5, value=val)
        val_cell.font = bold_font if "Net" in label or "Son" in label or "Hurda" in label else regular_font
        val_cell.number_format = '#,##0.00 "₺"'
        val_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Renklendirmeler
        if "Son Teklif" in label:
            val_cell.fill = total_fill
            lbl_cell.fill = total_fill
            val_cell.border = double_bottom_border
        elif "Hurda Değeri" in label:
            val_cell.fill = scrap_fill
            lbl_cell.fill = scrap_fill
            val_cell.border = thin_border
        else:
            val_cell.fill = summary_fill
            lbl_cell.fill = summary_fill
            
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = thin_border
            
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        
    # Kolon Genişliklerini Otomatik Ayarla
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ws.merged_cells:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(filename)


def export_to_pdf(teklif, kalemler, filename):
    """
    Teklifi resmi, şık ve Türkçe karakterleri destekleyen PDF dökümü olarak kaydeder.
    """
    doc = SimpleDocTemplate(
        filename,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Özel Stil Tanımlamaları
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName=FONT_BOLD_NAME,
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=5
    )
    
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#4B5563'),
        spaceAfter=15
    )
    
    label_style = ParagraphStyle(
        'MetaLabel',
        parent=styles['Normal'],
        fontName=FONT_BOLD_NAME,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#1F2937')
    )
    
    value_style = ParagraphStyle(
        'MetaValue',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#4B5563')
    )
    
    table_hdr_style = ParagraphStyle(
        'TableHdr',
        parent=styles['Normal'],
        fontName=FONT_BOLD_NAME,
        fontSize=9,
        leading=12,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#1F2937')
    )
    
    table_cell_right = ParagraphStyle(
        'TableCellRight',
        parent=table_cell_style,
        alignment=2 # Right
    )
    
    table_cell_center = ParagraphStyle(
        'TableCellCenter',
        parent=table_cell_style,
        alignment=1 # Center
    )
    
    story = []
    
    # ── 1. ÜST BAŞLIK ALANI ───────────────────────────────────────────────────
    story.append(Paragraph("ATÖLYE YÖNETİM SİSTEMİ", title_style))
    story.append(Paragraph("Resmi Teklif Formu ve Maliyet Analiz Raporu", subtitle_style))
    story.append(Spacer(1, 10))
    
    # ── 2. METADATA BİLGİ KARTLARI ────────────────────────────────────────────
    # Sol Sütun (Müşteri) & Sağ Sütun (Teklif Bilgileri)
    meta_data = [
        [
            Paragraph("Müşteri Firma:", label_style), Paragraph(teklif["firma_adi"] or "-", value_style),
            Paragraph("Teklif No:", label_style), Paragraph(teklif["teklif_no"], value_style)
        ],
        [
            Paragraph("Yetkili Telefon:", label_style), Paragraph(teklif["telefon"] or "-", value_style),
            Paragraph("Oluşturma Tarihi:", label_style), Paragraph(teklif["olusturma_tarihi"], value_style)
        ],
        [
            Paragraph("Yetkili E-posta:", label_style), Paragraph(teklif["mail"] or "-", value_style),
            Paragraph("Teslimat Tarihi:", label_style), Paragraph(teklif["teslim_tarihi"] or "-", value_style)
        ],
        [
            Paragraph("Teklif Başlığı:", label_style), Paragraph(teklif["baslik"] or "-", value_style),
            Paragraph("Teklif Durumu:", label_style), Paragraph(teklif["durum"], value_style)
        ]
    ]
    
    meta_table = Table(meta_data, colWidths=[90, 160, 90, 160])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#F3F4F6')),
    ]))
    
    story.append(meta_table)
    story.append(Spacer(1, 20))
    
    # ── 3. KALEMLER TABLOSU ───────────────────────────────────────────────────
    # Kolon Genişlikleri A4 Genişliğine (yaklaşık 515pt kullanılabilir alan) göre ayarlandı
    headers = [
        Paragraph("Malzeme Açıklaması", table_hdr_style),
        Paragraph("Miktar", table_hdr_style),
        Paragraph("Birim", table_hdr_style),
        Paragraph("Birim Fiyat", table_hdr_style),
        Paragraph("Toplam", table_hdr_style),
        Paragraph("Fire", table_hdr_style),
        Paragraph("Hurda", table_hdr_style)
    ]
    
    table_data = [headers]
    
    for k in kalemler:
        table_data.append([
            Paragraph(k["malzeme_adi"], table_cell_style),
            Paragraph(f"{k['miktar']:,.2f}", table_cell_right),
            Paragraph(k["birim"], table_cell_center),
            Paragraph(f"{k['birim_fiyat']:,.2f} TL", table_cell_right),
            Paragraph(f"{k['kalem_maliyeti']:,.2f} TL", table_cell_right),
            Paragraph(f"{k['fire_miktari']:,.2f}", table_cell_right),
            Paragraph(f"{k['tahmini_hurda_degeri']:,.2f} TL", table_cell_right)
        ])
        
    items_table = Table(table_data, colWidths=[155, 50, 40, 70, 75, 50, 75])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')])
    ]))
    
    story.append(items_table)
    story.append(Spacer(1, 15))
    
    # ── 4. ÖZET ALANI ─────────────────────────────────────────────────────────
    summary_data = [
        [Paragraph("Malzeme Toplam Maliyeti:", label_style), f"{teklif['malzeme_maliyeti']:,.2f} TL"],
        [Paragraph("Makine Toplam Maliyeti:", label_style), f"{teklif['makine_maliyeti']:,.2f} TL"],
        [Paragraph("Ek Operasyonel Giderler:", label_style), f"{teklif['ek_gider']:,.2f} TL"],
        [Paragraph("Net Üretim Maliyeti:", label_style), f"{teklif['net_maliyet']:,.2f} TL"],
        [Paragraph("Planlanan Kar Oranı / Tutarı:", label_style), f"{teklif['kar_tutari']:,.2f} TL"],
        [Paragraph("Manuel İndirim:", label_style), f"-{teklif['manuel_indirim']:,.2f} TL"],
        [Paragraph("Son Teklif Tutarı (KDV Hariç):", label_style), f"{teklif['son_tutar']:,.2f} TL"],
        [Paragraph("Geri Kazanılabilir Hurda Değeri:", label_style), f"{teklif['tahmini_hurda_degeri']:,.2f} TL"]
    ]
    
    summary_table_rows = []
    for label, val in summary_data:
        summary_table_rows.append([
            Paragraph("", table_cell_style), # Sol tarafı boş bırakmak için
            label,
            Paragraph(val, table_cell_right)
        ])
        
    summary_table = Table(summary_table_rows, colWidths=[240, 180, 95])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW', (1,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('BACKGROUND', (1,6), (2,6), colors.HexColor('#D1FAE5')), # Son Toplam Yeşil Arka Plan
        ('BACKGROUND', (1,7), (2,7), colors.HexColor('#FEF3C7')), # Hurda Sarı Arka Plan
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 30))
    
    # ── 5. ALT BİLGİ NOTU ─────────────────────────────────────────────────────
    not_style = ParagraphStyle(
        'FooterNote',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#9CA3AF'),
        alignment=1 # Center
    )
    story.append(Paragraph("Bu belge Atölye Yönetim ERP Sistemi tarafından otomatik olarak oluşturulmuştur.", not_style))
    story.append(Paragraph(f"Oluşturulma Zamanı: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}", not_style))
    
    # PDF İnşa Et
    doc.build(story)

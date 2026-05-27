import os
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Türkçe Karakter Destekli Arial Yazı Tipi Yükleme
FONT_NAME = "Helvetica"
FONT_BOLD_NAME = "Helvetica-Bold"
arial_path = "C:\\Windows\\Fonts\\arial.ttf"
arial_bold_path = "C:\\Windows\\Fonts\\arialbd.ttf"

if os.path.exists(arial_path) and os.path.exists(arial_bold_path):
    try:
        pdfmetrics.registerFont(TTFont("Arial", arial_path))
        pdfmetrics.registerFont(TTFont("Arial-Bold", arial_bold_path))
        FONT_NAME = "Arial"
        FONT_BOLD_NAME = "Arial-Bold"
    except Exception:
        pass

def export_rapor_to_excel(data, filename):
    wb = openpyxl.Workbook()
    
    # ── 1. ÖZET SAYFASI ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Özet ve KPI"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1E3A8A")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="6B7280")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True, color="111827")
    regular_font = Font(name=font_family, size=11, color="374151")
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    kpi_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    # Başlık
    ws["A1"] = "ATÖLYE YÖNETİM SİSTEMİ - ANALİZ RAPORU"
    ws["A1"].font = title_font
    ws["A2"] = f"Raporlama Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = subtitle_font
    
    # Filtreler
    ws["A4"] = "UYGULANAN FİLTRELER"
    ws["A4"].font = bold_font
    filters = data["filters"]
    ws["A5"] = "Başlangıç Tarihi:"
    ws["B5"] = filters.get("tarih_bas", "-")
    ws["A6"] = "Bitiş Tarihi:"
    ws["B6"] = filters.get("tarih_bit", "-")
    ws["A7"] = "Müşteri:"
    ws["B7"] = filters.get("musteri", "-")
    ws["A8"] = "İşlem Türü:"
    ws["B8"] = filters.get("islem", "-")
    ws["A9"] = "Teklif Durumu:"
    ws["B9"] = filters.get("durum", "-")
    for r in range(5, 10):
        ws.cell(row=r, column=1).font = bold_font
        ws.cell(row=r, column=2).font = regular_font
        
    # KPI Kartları
    ws["D4"] = "KPI ÖZETLERİ"
    ws["D4"].font = bold_font
    kpi = data["kpi"]
    kpi_items = [
        ("Toplam Teklif Tutarı", kpi.get("toplam_tutar", 0.0), '₺'),
        ("Toplam Net Maliyet", kpi.get("toplam_maliyet", 0.0), '₺'),
        ("Toplam Kar", kpi.get("toplam_kar", 0.0), '₺'),
        ("Ortalama Kar Oranı", kpi.get("ort_kar_orani", 0.0), '%'),
        ("Tahmini Hurda Değeri", kpi.get("tahmini_hurda", 0.0), '₺')
    ]
    r_idx = 5
    for label, val, unit in kpi_items:
        ws.cell(row=r_idx, column=4, value=label).font = bold_font
        c_val = ws.cell(row=r_idx, column=5, value=val)
        c_val.font = bold_font
        c_val.fill = kpi_fill
        if unit == '₺':
            c_val.number_format = '#,##0.00 "₺"'
        elif unit == '%':
            c_val.number_format = '0.0 "%"'
        r_idx += 1

    # ── 2. MÜŞTERİ BAZLI RAPOR SEKMESİ ─────────────────────────────────────────
    ws_cust = wb.create_sheet(title="Müşteri Analizi")
    ws_cust.views.sheetView[0].showGridLines = True
    ws_cust["A1"] = "Müşteri Bazlı Rapor"
    ws_cust["A1"].font = title_font
    headers_cust = ["Müşteri Firma", "Teklif Adedi", "Toplam Tutar", "Toplam Maliyet", "Toplam Kar"]
    for col, h in enumerate(headers_cust, 1):
        cell = ws_cust.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    r_idx = 4
    for row in data["musteri_raporu"]:
        ws_cust.cell(row=r_idx, column=1, value=row["firma_adi"]).font = regular_font
        ws_cust.cell(row=r_idx, column=2, value=row["teklif_adedi"]).font = regular_font
        
        c_tot = ws_cust.cell(row=r_idx, column=3, value=row["toplam_tutar"])
        c_tot.font = regular_font
        c_tot.number_format = '#,##0.00 "₺"'
        
        c_cost = ws_cust.cell(row=r_idx, column=4, value=row["toplam_maliyet"])
        c_cost.font = regular_font
        c_cost.number_format = '#,##0.00 "₺"'
        
        c_profit = ws_cust.cell(row=r_idx, column=5, value=row["toplam_kar"])
        c_profit.font = regular_font
        c_profit.number_format = '#,##0.00 "₺"'
        r_idx += 1
        
    # ── 3. İŞLEM BAZLI RAPOR SEKMESİ ──────────────────────────────────────────
    ws_op = wb.create_sheet(title="İşlem Analizi")
    ws_op.views.sheetView[0].showGridLines = True
    ws_op["A1"] = "İşlem Bazlı Rapor"
    ws_op["A1"].font = title_font
    headers_op = ["İşlem / Makine", "Kullanım Adedi", "Toplam Süre (Saat)", "Toplam Maliyet Katkısı"]
    for col, h in enumerate(headers_op, 1):
        cell = ws_op.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    r_idx = 4
    for row in data["islem_raporu"]:
        ws_op.cell(row=r_idx, column=1, value=row["islem_adi"]).font = regular_font
        ws_op.cell(row=r_idx, column=2, value=row["kullanim_sayisi"]).font = regular_font
        
        c_dur = ws_op.cell(row=r_idx, column=3, value=row["toplam_sure"])
        c_dur.font = regular_font
        c_dur.number_format = '#,##0.0'
        
        c_cost = ws_op.cell(row=r_idx, column=4, value=row["toplam_maliyet"])
        c_cost.font = regular_font
        c_cost.number_format = '#,##0.00 "₺"'
        r_idx += 1

    # ── 4. MALZEME BAZLI RAPOR SEKMESİ ────────────────────────────────────────
    ws_mat = wb.create_sheet(title="Malzeme Analizi")
    ws_mat.views.sheetView[0].showGridLines = True
    ws_mat["A1"] = "Malzeme Bazlı Rapor"
    ws_mat["A1"].font = title_font
    headers_mat = ["Malzeme Adı", "Kullanılan Miktar", "Birim", "Toplam Maliyet Katkısı"]
    for col, h in enumerate(headers_mat, 1):
        cell = ws_mat.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    r_idx = 4
    for row in data["malzeme_raporu"]:
        ws_mat.cell(row=r_idx, column=1, value=row["malzeme_adi"]).font = regular_font
        
        c_qty = ws_mat.cell(row=r_idx, column=2, value=row["toplam_miktar"])
        c_qty.font = regular_font
        c_qty.number_format = '#,##0.00'
        
        ws_mat.cell(row=r_idx, column=3, value=row["birim"]).font = regular_font
        
        c_cost = ws_mat.cell(row=r_idx, column=4, value=row["toplam_maliyet"])
        c_cost.font = regular_font
        c_cost.number_format = '#,##0.00 "₺"'
        r_idx += 1
        
    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb.save(filename)


def export_rapor_to_pdf(data, filename):
    doc = SimpleDocTemplate(
        filename,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'RaporTitle',
        parent=styles['Normal'],
        fontName=FONT_BOLD_NAME,
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'RaporSub',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#4B5563'),
        spaceAfter=15
    )
    
    section_style = ParagraphStyle(
        'RaporSec',
        parent=styles['Normal'],
        fontName=FONT_BOLD_NAME,
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#1F2937'),
        spaceBefore=12,
        spaceAfter=6
    )
    
    label_style = ParagraphStyle(
        'RaporLabel',
        parent=styles['Normal'],
        fontName=FONT_BOLD_NAME,
        fontSize=9,
        leading=12
    )
    
    value_style = ParagraphStyle(
        'RaporValue',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=9,
        leading=12
    )
    
    table_hdr_style = ParagraphStyle(
        'TableHdrR',
        parent=styles['Normal'],
        fontName=FONT_BOLD_NAME,
        fontSize=9,
        leading=12,
        textColor=colors.white,
        alignment=1
    )
    
    table_cell_style = ParagraphStyle(
        'TableCellR',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=9,
        leading=12
    )
    
    table_cell_right = ParagraphStyle(
        'TableCellRightR',
        parent=table_cell_style,
        alignment=2
    )
    
    table_cell_center = ParagraphStyle(
        'TableCellCenterR',
        parent=table_cell_style,
        alignment=1
    )
    
    story = []
    
    # ── 1. BAŞLIK ALANI ───────────────────────────────────────────────────────
    story.append(Paragraph("ATÖLYE YÖNETİM SİSTEMİ - ANALİZ RAPORU", title_style))
    story.append(Paragraph(f"Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}", subtitle_style))
    story.append(Spacer(1, 8))
    
    # ── 2. UYGULANAN FİLTRELER VE KPI BLOKLARI ──────────────────────────────────
    filters = data["filters"]
    kpi = data["kpi"]
    
    left_meta = [
        [Paragraph("Başlangıç Tarihi:", label_style), Paragraph(filters.get("tarih_bas", "-"), value_style)],
        [Paragraph("Bitiş Tarihi:", label_style), Paragraph(filters.get("tarih_bit", "-"), value_style)],
        [Paragraph("Müşteri Firma:", label_style), Paragraph(filters.get("musteri", "-"), value_style)],
        [Paragraph("İşlem Türü:", label_style), Paragraph(filters.get("islem", "-"), value_style)],
        [Paragraph("Teklif Durumu:", label_style), Paragraph(filters.get("durum", "-"), value_style)]
    ]
    
    right_kpis = [
        [Paragraph("Toplam Teklif Tutarı:", label_style), Paragraph(f"{kpi.get('toplam_tutar', 0.0):,.2f} TL", value_style)],
        [Paragraph("Toplam Net Maliyet:", label_style), Paragraph(f"{kpi.get('toplam_maliyet', 0.0):,.2f} TL", value_style)],
        [Paragraph("Toplam Kar:", label_style), Paragraph(f"{kpi.get('toplam_kar', 0.0):,.2f} TL", value_style)],
        [Paragraph("Ortalama Kar Oranı:", label_style), Paragraph(f"% {kpi.get('ort_kar_orani', 0.0):,.1f}", value_style)],
        [Paragraph("Tahmini Hurda Değeri:", label_style), Paragraph(f"{kpi.get('tahmini_hurda', 0.0):,.2f} TL", value_style)]
    ]
    
    meta_table_data = [
        [
            Table(left_meta, colWidths=[100, 150]),
            Spacer(1, 1),
            Table(right_kpis, colWidths=[130, 120])
        ]
    ]
    
    meta_table = Table(meta_table_data, colWidths=[250, 30, 250])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    
    story.append(meta_table)
    story.append(Spacer(1, 15))
    
    # ── 3. MÜŞTERİ BAZLI ANALİZ TABLOSU ────────────────────────────────────────
    story.append(Paragraph("Müşteri Bazlı Rapor", section_style))
    m_headers = [
        Paragraph("Müşteri Firma", table_hdr_style),
        Paragraph("Teklif Adedi", table_hdr_style),
        Paragraph("Toplam Tutar", table_hdr_style),
        Paragraph("Toplam Maliyet", table_hdr_style),
        Paragraph("Toplam Kar", table_hdr_style)
    ]
    
    cust_data = [m_headers]
    for row in data["musteri_raporu"]:
        cust_data.append([
            Paragraph(row["firma_adi"], table_cell_style),
            Paragraph(str(row["teklif_adedi"]), table_cell_center),
            Paragraph(f"{row['toplam_tutar']:,.2f} TL", table_cell_right),
            Paragraph(f"{row['toplam_maliyet']:,.2f} TL", table_cell_right),
            Paragraph(f"{row['toplam_kar']:,.2f} TL", table_cell_right)
        ])
        
    cust_table = Table(cust_data, colWidths=[175, 70, 95, 95, 95])
    cust_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')])
    ]))
    story.append(cust_table)
    story.append(Spacer(1, 12))
    
    # ── 4. İŞLEM BAZLI ANALİZ TABLOSU ──────────────────────────────────────────
    story.append(Paragraph("İşlem Bazlı Rapor", section_style))
    op_headers = [
        Paragraph("İşlem / Makine Adı", table_hdr_style),
        Paragraph("Kullanım Sayısı", table_hdr_style),
        Paragraph("Toplam Süre (Saat)", table_hdr_style),
        Paragraph("Toplam Maliyet Katkısı", table_hdr_style)
    ]
    
    op_data = [op_headers]
    for row in data["islem_raporu"]:
        op_data.append([
            Paragraph(row["islem_adi"], table_cell_style),
            Paragraph(str(row["kullanim_sayisi"]), table_cell_center),
            Paragraph(f"{row['toplam_sure']:,.1f}", table_cell_right),
            Paragraph(f"{row['toplam_maliyet']:,.2f} TL", table_cell_right)
        ])
        
    op_table = Table(op_data, colWidths=[200, 90, 110, 130])
    op_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')])
    ]))
    story.append(op_table)
    story.append(Spacer(1, 12))
    
    # ── 5. MALZEME BAZLI ANALİZ TABLOSU ────────────────────────────────────────
    story.append(Paragraph("Malzeme Bazlı Rapor", section_style))
    mat_headers = [
        Paragraph("Malzeme Adı", table_hdr_style),
        Paragraph("Toplam Kullanılan Miktar", table_hdr_style),
        Paragraph("Birim", table_hdr_style),
        Paragraph("Toplam Maliyet Katkısı", table_hdr_style)
    ]
    
    mat_data = [mat_headers]
    for row in data["malzeme_raporu"]:
        mat_data.append([
            Paragraph(row["malzeme_adi"], table_cell_style),
            Paragraph(f"{row['toplam_miktar']:,.2f}", table_cell_right),
            Paragraph(row["birim"], table_cell_center),
            Paragraph(f"{row['toplam_maliyet']:,.2f} TL", table_cell_right)
        ])
        
    mat_table = Table(mat_data, colWidths=[210, 110, 80, 130])
    mat_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')])
    ]))
    story.append(mat_table)
    story.append(Spacer(1, 20))
    
    # ── 6. ALT BİLGİ ──────────────────────────────────────────────────────────
    not_style = ParagraphStyle(
        'FooterNoteR',
        parent=styles['Normal'],
        fontName=FONT_NAME,
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#9CA3AF'),
        alignment=1
    )
    story.append(Paragraph("Bu analiz dökümü Atölye Yönetim ERP Sistemi Raporlama Modülü tarafından oluşturulmuştur.", not_style))
    
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    doc.build(story)
